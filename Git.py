import requests
import json
import time
import sys
from time import mktime
from jsonpath_ng import parse
from datetime import datetime
from openpyxl import load_workbook
from datetime import timedelta  
import calendar



wb = load_workbook('Analytics.xlsx')
default_hdr = {}
commit_hdr = {'Accept':'application/vnd.github.cloak-preview'}
def issues(owner, repo, start, end, row_index):
    date_range = start+".."+end
    write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " created:"+date_range, row_index, 4, default_hdr)
    #write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo +" is:closed created:"+date_range, row_index, 5, default_hdr)
    #write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:open created:"+date_range, row_index, 6, default_hdr)
    write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:closed closed:"+date_range, row_index, 7,default_hdr)

        
def pr(owner, repo, start, end, row_index):
    date_range = start+".."+end
    write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:pr is:closed created:"+date_range, row_index, 9,default_hdr)
    write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:pr linked:issue created:"+date_range, row_index,  10,default_hdr)
    #write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:pr is:merged created:"+date_range, row_index, 11,default_hdr)
    write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:pr interactions:0 created:"+date_range, row_index, 12,default_hdr)
    write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:pr interactions:1..20 created:"+date_range, row_index, 13,default_hdr)
    #write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:pr interactions:11..20 created:"+date_range,row_index,  14,default_hdr)
    write_to_excel("https://api.github.com/search/issues?q=" + owner + "/" + repo + " is:pr interactions:21..* created:"+date_range, row_index, 15,default_hdr)


def commits(owner, repo, start, end, row_index):
    date_range = start+".."+end
    #write_to_excel("https://api.github.com/search/commits?q=" + owner + "/" + repo + " merge:false author-date:" + date_range, row_index, 16, commit_hdr)
    write_to_excel("https://api.github.com/search/commits?q=" + owner + "/" + repo + " merge:true author-date:"+date_range, row_index, 17,commit_hdr)
    write_to_excel("https://api.github.com/search/commits?q=" + owner + "/" + repo + " author-date:" +date_range, row_index, 18,commit_hdr)


def write_to_excel(url, row_index, column_index, hdr):
    ws = wb['Data-PR-Issues-Commits']
    print(url)
    rate_limit()
    response = requests.get(url,headers=hdr)

    while (response.status_code!=200):
        print("Error getting data for " + url + " " + str(response.status_code))
        time.sleep(60)
        response = requests.get(url,headers=hdr)

    issues = json.loads(response.content.decode('utf-8'))
    cc = ws.cell(row_index, column_index)
    cc.value = issues['total_count']

        

def write_headers(owner, repo, row_index, end_date):
    global wb
    ws=wb['Data-PR-Issues-Commits']
    c1 = ws.cell(row=row_index, column=1)
    c1.value = owner
    c2 = ws.cell(row=row_index, column=2)
    c2.value = repo
    c3 = ws.cell(row=row_index, column=3)
    c3.value = end_date
    
                

def rate_limit(dbg=False):
    url = "https://api.github.com/rate_limit"
    response = requests.get(url,headers=default_hdr)
    rate_json = json.loads(response.text)
    if response.status_code == 200:
        json_expression=parse('$.resources.search.reset')
        match=json_expression.find(rate_json)   
        reset_time = time.localtime(match[0].value)
        dt_string = time.strftime("%m/%d/%Y %H:%M:%S", reset_time)
        now = datetime.now()
            
        json_expression_remaining=parse('$.resources.search.remaining')
        match=json_expression_remaining.find(rate_json) 
        if dbg:
            print(rate_json)
            print("reset date and time =", dt_string)
            print("now =", now)
            print('Remaining' + str(match[0].value))
    
        if (match[0].value == 1):
            dt = datetime.fromtimestamp(mktime(reset_time))
            if (dt > now):
                print ('Sleeping ' + str((dt-now).total_seconds()))
                time.sleep((dt-now).total_seconds()+1)
    else:
        print(response.text)


def write_contributors(owner, repo, row_index):
    global wb
    ws = wb['Contributors-data']
    print ('Working on write_contributors()')
    
    url = "https://api.github.com/repos/"+owner+"/"+repo+"/stats/contributors"
    response = requests.get(url,headers=default_hdr)
    
    if (response.status_code == 200):
        conjson = json.loads(response.text)
        #with open('cont.json') as json_file:
        #    conjson = json.load(json_file)
        json_exp = parse('$[*]')
        additions_tot=0
        deletions_tot=0
        commits_tot=0
        lists= [match.value for match in json_exp.find(conjson)]
        for l in lists:
            if (type(l) == type(dict())):
                for (k, v) in l.items():
                    if (k=="author"):
                        for a,a2 in v.items():
                            if (a=="login"):
                                print(a2)
                                print(additions_tot)
                                print(deletions_tot)
                                print(commits_tot)
                                c1 = ws.cell(row=row_index, column=1)
                                c1.value = owner
                                c2 = ws.cell(row=row_index, column=2)
                                c2.value = repo
                                cc = ws.cell(row_index, column=3)
                                cc.value = a2
                                c4 = ws.cell(row_index, column=4)
                                c4.value = additions_tot
                                c5 = ws.cell(row_index, column=5)
                                c5.value = deletions_tot
                                c6 = ws.cell(row_index, column=6)
                                c6.value = commits_tot
                                
                                row_index = row_index+1
                                additions_tot=0
                                deletions_tot=0
                                commits_tot=0
                    elif (k == "weeks"):
                        for v1 in v:
                            additions_tot=additions_tot+v1["a"]
                            deletions_tot=deletions_tot+v1["d"]
                            commits_tot=commits_tot+v1["c"]
    else:
        print("Error getting data for contributors" + url + " "+response.text)    
    
    return row_index
    
    
    
    

def test_contributors():
    global repos
    #global wb
    #row_index=2
 
    #for repo in repos :
    url = "https://api.github.com/repos/"+"tensorflow/tensorflow"+"/stats/contributors"
    response = requests.get(url,headers=default_hdr)
    
    if (response.status_code == 200):
        #conjson = json.loads(response.text)
        with open('cont.json') as json_file:
            conjson = json.load(json_file)
        json_exp = parse('$[*]')
        additions_tot=0
        deletions_tot=0
        commits_tot=0
        lists= [match.value for match in json_exp.find(conjson)]
        for l in lists:
            if (type(l) == type(dict())):
                for (k, v) in l.items():
                    if (k=="author"):
                        for a,a2 in v.items():
                            if (a=="login"):
                                print(a2)
                                print(additions_tot)
                                print(deletions_tot)
                                print(commits_tot)
                                additions_tot=0
                                deletions_tot=0
                                commits_tot=0
                    elif (k == "weeks"):
                        for v1 in v:
                            print(v1["w"])
                            additions_tot=additions_tot+v1["a"]
                            deletions_tot=deletions_tot+v1["d"]
                            commits_tot=commits_tot+v1["c"]
                            #print("a "+str(v1["a"]))
                            #print(v1["d"])
                            #print(v1["c"])
    else:
        print("Error getting data for contributors" + url + " "+response.text)

def write_commit_activity(owner, repo, row_index):
    global wb
    start_row_index=row_index
    ws = wb['Commit-Activity']
    
    print('Working on write_commit_activity')
 
    url = "https://api.github.com/repos/"+owner+"/"+repo+"/stats/commit_activity"
    print(url)
    response = requests.get(url,headers=default_hdr)
    
    if (response.status_code == 200):
        conjson = json.loads(response.text)
        row_index=start_row_index
        week_exp = parse('$.[*].week')
        for match in week_exp.find(conjson):
            c1 = ws.cell(row=row_index, column=1)
            c1.value = owner
            c2 = ws.cell(row=row_index, column=2)
            c2.value = owner
            week_date = time.localtime(match.value)
            dt_string = time.strftime("%m/%d/%Y", week_date)
            #print(f'{match.value}')
            c3 = ws.cell(row_index, column=3)
            c3.value=dt_string
            c4 = ws.cell(row_index, column=4)
            c4.value = time.strftime("%Y", week_date)
            row_index = row_index+1
        
        row_index=start_row_index    
        total_exp = parse('$.[*].total')
        for match in total_exp.find(conjson):
            #print(f'{match.value}')
            c5 = ws.cell(row_index, column=5)
            c5.value=f'{match.value}'
            row_index=row_index+1
            
        start_row_index=row_index

    else:
        print ("Commit_activity Call Failed " + response.text)
        
    return row_index
            
def write_code_frequency(owner, repo, row_index):
    global wb
 
    print('Working on write_code_frequency')
    url = "https://api.github.com/repos/"+owner+"/"+repo+"/stats/code_frequency"
    print(url)
    response = requests.get(url,headers=default_hdr)
    
    if (response.status_code == 200):
        conjson = json.loads(response.text)
        ws = wb['Code-Frequency']
        json_exp = parse('$.[*]')
        #for match in json_exp.find(conjson):
        week_lists= [match.value for match in json_exp.find(conjson)]
        for week_list in week_lists:
            c1 = ws.cell(row=row_index, column=1)
            c1.value = owner
            c2 = ws.cell(row=row_index, column=2)
            c2.value = repo
            c3 = ws.cell(row_index, column=3)
            week_date = time.localtime(week_list[0])
            dt_string = time.strftime("%m/%d/%Y", week_date)
            print (repo + " " + dt_string)
            c3.value = dt_string
            c4 = ws.cell(row_index, column=4)
            c4.value = week_list[1]
            c5 = ws.cell(row_index, column=5)
            c5.value = week_list[2]
            c6 = ws.cell(row_index, column=6)
            c6.value = time.strftime("%Y", week_date)
            row_index = row_index+1
    else:
        print ("Code Frequency Call Failled " + response.text)
    
    return row_index



def issues_pr_commits():
    global wb
    ws = wb["input"]
    input_row=ws.cell(row=1, column= 7).value
    tot = ws.cell(row=2, column=7).value
    print(datetime.now())
    row_index = ws.cell(row=3, column= 7).value
    
    while (input_row <= tot):
        owner=ws.cell(input_row, 1).value
        print(owner)
        repo=ws.cell(input_row, 2).value
        print(repo)
        start_date = ws.cell(input_row, 3).value
        end_date= ws.cell(input_row, 4).value
        while (start_date < end_date):
            dt_string1 = start_date.strftime("%Y-%m-%d")
            #start_date=start_date+timedelta(days=interval)
            days_in_month = calendar.monthrange(start_date.year, start_date.month)[1]
            dt_string2 = (start_date + timedelta(days=days_in_month-1)).strftime("%Y-%m-%d")
            
            write_headers(owner, repo, row_index, start_date)
            issues(owner, repo, dt_string1, dt_string2, row_index)
            pr(owner, repo, dt_string1, dt_string2, row_index)
            commits(owner, repo, dt_string1, dt_string2, row_index)
           
            row_index=row_index+1
            
            start_date = start_date + timedelta(days=days_in_month)
        
        input_row=input_row+1
        wb.save("Analytics.xlsx")


def code_metrics():
    global wb
    ws = wb["input"]
    input_row=ws.cell(row=1, column=7).value
    tot = ws.cell(row=2, column=7).value
    print(datetime.now())
    row_index=2

    while (input_row <= tot):
        owner=ws.cell(input_row, 1).value
        print(owner)
        repo=ws.cell(input_row, 2).value
        row_index=write_code_frequency(owner, repo, row_index)
        input_row=input_row+1

    row_index=2
    input_row=ws.cell(row=1, column=7).value
    while (input_row <= tot):
        owner=ws.cell(input_row, 1).value
        print(owner)
        repo=ws.cell(input_row, 2).value
        row_index=write_commit_activity(owner, repo, row_index)
        input_row=input_row+1
        
    row_index=2
    input_row=ws.cell(row=1, column=7).value
    while (input_row <= tot):
        owner=ws.cell(input_row, 1).value
        print(owner)
        repo=ws.cell(input_row, 2).value
        row_index=write_contributors(owner, repo, row_index)
        input_row=input_row+1
        
    
    
    # Print the contents
    
def main():
    # print command line arguments
    global default_hdr
    global commit_hdr
    for arg in sys.argv[1:]:
        default_hdr = {'Authorization': 'Token '+arg+''}
        commit_hdr = {'Authorization': 'Token '+arg+'','Accept':'application/vnd.github.cloak-preview'}
        print(default_hdr)
        print(commit_hdr)
        issues_pr_commits()
        code_metrics()
    
    
if __name__ == "__main__":
    main()    
                 
wb.save("Analytics.xlsx")

